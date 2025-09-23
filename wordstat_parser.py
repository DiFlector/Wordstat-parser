#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Программа для парсинга данных из Яндекс Вордстат
Автор: DiFlector
"""

import requests
import time
import re
import os
import sys
import shutil
from urllib.parse import quote, urlencode
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class WordstatParser:
    """Класс для парсинга данных из Яндекс Вордстат"""
    
    def __init__(self, use_selenium=True):
        """
        Инициализация парсера
        
        Args:
            use_selenium (bool): Использовать ли Selenium (рекомендуется для Яндекса)
        """
        self.use_selenium = use_selenium
        self.base_url = "https://wordstat.yandex.ru/"
        self.driver = None
        self.is_authorized = False  # Флаг авторизации
        
        if use_selenium:
            self._init_selenium()
    
    def _init_selenium(self):
        """Инициализация Selenium WebDriver"""
        print("🚀 Инициализация Selenium WebDriver...")
        
        chrome_options = Options()
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--disable-web-security")
        chrome_options.add_argument("--allow-running-insecure-content")
        chrome_options.add_argument("--disable-extensions")
        # Добавляем User-Agent для имитации обычного браузера
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36")
        
        # Пробуем несколько способов инициализации
        methods = [
            ("webdriver-manager", self._init_with_webdriver_manager),
            ("системный chromedriver", self._init_with_system_chrome),
            ("локальный chromedriver", self._init_with_local_chrome),
            ("Firefox WebDriver", self._init_with_firefox)
        ]
        
        for method_name, method_func in methods:
            try:
                print(f"  Пробуем: {method_name}...")
                if method_func(chrome_options):
                    print(f"✓ Selenium WebDriver инициализирован через {method_name}")
                    return
            except Exception as e:
                print(f"  ✗ {method_name} не работает: {e}")
                continue
        
        print("✗ Не удалось инициализировать ни один WebDriver")
        print("  Переключаемся на режим без Selenium (менее надежно)")
        self.use_selenium = False
    
    def _init_with_webdriver_manager(self, chrome_options):
        """Инициализация через webdriver-manager"""
        # Очищаем кэш webdriver-manager
        try:
#            cache_path = os.path.join(os.path.expanduser("~"), ".wdm")
#            if os.path.exists(cache_path):
#                shutil.rmtree(cache_path)
                print("  Очищен кэш webdriver-manager")
        except:
            pass
            
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        return True
    
    def _init_with_system_chrome(self, chrome_options):
        """Инициализация с системным chromedriver"""
        service = Service()  # Использует chromedriver из PATH
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        return True
    
    def _init_with_local_chrome(self, chrome_options):
        """Инициализация с локальным chromedriver"""
        local_paths = [
            "chromedriver.exe",
            "./chromedriver.exe",
            "../chromedriver.exe"
        ]
        
        for path in local_paths:
            if os.path.exists(path):
                service = Service(path)
                self.driver = webdriver.Chrome(service=service, options=chrome_options)
                return True
        return False
    
    def _init_with_firefox(self, options_unused):
        """Инициализация с Firefox как запасной вариант"""
        try:
            from selenium.webdriver.firefox.options import Options as FirefoxOptions
            from selenium.webdriver.firefox.service import Service as FirefoxService
            from webdriver_manager.firefox import GeckoDriverManager
            
            firefox_options = FirefoxOptions()
            firefox_options.add_argument("--width=1920")
            firefox_options.add_argument("--height=1080")
            firefox_options.add_argument("--disable-gpu")
            
            service = FirefoxService(GeckoDriverManager().install())
            self.driver = webdriver.Firefox(service=service, options=firefox_options)
            return True
        except Exception:
            return False
    
    def authorize_wordstat(self):
        """
        Авторизация в Яндекс Вордстат с интерактивным ожиданием
        
        Returns:
            bool: True если авторизация прошла успешно
        """
        if not self.driver:
            print("❌ WebDriver не инициализирован")
            return False
        
        if self.is_authorized:
            print("✅ Уже авторизован в Вордстат")
            return True
        
        try:
            print("🔐 Начинаем процедуру авторизации в Яндекс Вордстат...")
            print("=" * 60)
            
            # Открываем Вордстат с тестовым запросом
            test_query = "тест"
            test_url = self.build_wordstat_url(test_query)
            
            print(f"🌐 Открываем Вордстат: {test_url}")
            self.driver.get(test_url)
            
            # Ждем немного для загрузки страницы
            time.sleep(3)
            
            # Проверяем, требуется ли авторизация
            page_source = self.driver.page_source.lower()
            page_title = self.driver.title.lower()
            
            # Если уже авторизован
            if "вордстат" in page_title and ("войти" not in page_source or "login" not in page_source):
                print("✅ Авторизация не требуется - уже вошли в систему")
                self.is_authorized = True
                return True
            
            # Если требуется авторизация
            print("⚠️  Требуется авторизация в Яндекс аккаунт")
            print("")
            print("🔍 ИНСТРУКЦИЯ:")
            print("1. В открывшемся браузере войдите в ваш Яндекс аккаунт")
            print("2. Дождитесь загрузки страницы Вордстат")
            print("3. Убедитесь, что видите интерфейс Вордстат")
            print("4. Вернитесь в эту программу")
            print("")
            print("⏳ У вас есть 10 секунд для входа...")
            
            # Интерактивное ожидание с обратным отсчетом
            for i in range(10, 0, -1):
                print(f"   Осталось: {i} сек", end="\r")
                time.sleep(1)
            
            print("\n")
            print("🔍 Проверяем авторизацию...")
            
            # Обновляем страницу для проверки
            self.driver.refresh()
            time.sleep(3)
            
            # Проверяем результат авторизации
            current_url = self.driver.current_url
            page_title = self.driver.title
            page_source = self.driver.page_source.lower()
            
            print(f"   Текущий URL: {current_url}")
            print(f"   Заголовок: {page_title}")
            
            # Критерии успешной авторизации
            success_indicators = [
                "вордстат" in page_title.lower(),
                "wordstat" in current_url,
                "запросов" in page_source,
                "частота" in page_source
            ]
            
            failure_indicators = [
                "войти" in page_source,
                "login" in page_source,
                "авторизация" in page_source,
                "passport.yandex" in current_url
            ]
            
            success_count = sum(success_indicators)
            failure_count = sum(failure_indicators)
            
            if success_count >= 2 and failure_count == 0:
                print("✅ Авторизация успешна!")
                print("   Переходим к парсингу с ускоренными запросами")
                self.is_authorized = True
                return True
            else:
                print("❌ Авторизация не удалась")
                print("   Программа будет работать с большими задержками")
                print(f"   Индикаторы успеха: {success_count}/4")
                print(f"   Индикаторы неудачи: {failure_count}")
                
                # Даем еще один шанс
                choice = input("\n🔄 Попробовать еще раз? (y/n): ").lower()
                if choice in ['y', 'yes', 'да', 'д']:
                    return self.authorize_wordstat()
                
                return False
                
        except Exception as e:
            print(f"❌ Ошибка авторизации: {e}")
            return False
    
    def format_query(self, query, query_type="base"):
        """
        Форматирование запроса в зависимости от типа
        
        Args:
            query (str): Исходный запрос
            query_type (str): Тип запроса ("base", "exact", "precise")
            
        Returns:
            str: Отформатированный запрос
        """
        query = query.strip()
        
        if query_type == "base":
            # Базовый запрос - как есть
            return query
        elif query_type == "exact":
            # Точный запрос - в кавычках
            return f'"{query}"'
        elif query_type == "precise":
            # Уточненный запрос - в кавычках с ! перед каждым словом
            words = query.split()
            precise_words = [f"!{word}" for word in words]
            return f'"{" ".join(precise_words)}"'
        else:
            return query
    
    def build_wordstat_url(self, query):
        """
        Построение URL для Яндекс Вордстат
        
        Args:
            query (str): Поисковый запрос
            
        Returns:
            str: URL для запроса
        """
        params = {
            'region': 'all',
            'view': 'table',
            'words': query
        }
        return f"{self.base_url}?{urlencode(params)}"
    
    def parse_frequency_selenium(self, query):
        """
        Парсинг частоты запроса с использованием Selenium
        
        Args:
            query (str): Поисковый запрос
            
        Returns:
            int or None: Частота запроса или None в случае ошибки
        """
        if not self.driver:
            return None
            
        try:
            url = self.build_wordstat_url(query)
            print(f"  Запрос: {query}")
            print(f"  URL: {url}")
            
            self.driver.get(url)
            
            # Ждем загрузки страницы (короче если авторизован)
            if self.is_authorized:
                time.sleep(0.5)  # Быстро если авторизован
            else:
                time.sleep(5)    # Медленно если не авторизован
            
            # Ищем элементы с частотой запроса в Яндекс Вордстат
            frequency_selectors = [
                # Новые селекторы для актуального интерфейса Яндекс Вордстат
                '.wordstat__content-preview-text_last',
                '.wordstat__content-preview-text',
                '.wordstat__number',
                '.wordstat-number',
                
                # Селекторы для общего числа запросов
                '[class*="wordstat__"]',
                '[class*="preview-text"]',
                
                # Старые селекторы (на всякий случай)
                '.wordstat-table__row:first-child .wordstat-table__cell:nth-child(2)',
                '.table__row:first-child .table__cell:nth-child(2)',
                '[data-testid="frequency"]',
                '.frequency',
                '.stat-value'
            ]
            
            frequency = None
            
            # Метод 1: Поиск по CSS селекторам
            for selector in frequency_selectors:
                try:
                    elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    for element in elements:
                        text = element.text.strip()
                        print(f"    Найден элемент '{selector}': {text}")
                        
                        # Ищем числа в формате "за дата – дата: ЧИСЛО"
                        frequency_match = re.search(r':\s*(\d{1,3}(?:\s\d{3})*)', text)
                        if frequency_match:
                            frequency = int(frequency_match.group(1).replace(' ', ''))
                            print(f"    Извлечена частота из паттерна ': ЧИСЛО': {frequency}")
                            break
                        
                        # Ищем числа в общем тексте
                        numbers = re.findall(r'\b(\d{1,3}(?:\s\d{3})*)\b', text)
                        if numbers:
                            frequency = int(numbers[-1].replace(' ', ''))  # Берем последнее число
                            print(f"    Извлечена частота из чисел: {frequency}")
                            break
                    
                    if frequency:
                        break
                except Exception as e:
                    print(f"    Ошибка с селектором '{selector}': {e}")
                    continue
            
            # Метод 2: Поиск в заголовках и подзаголовках
            if frequency is None:
                try:
                    headings = self.driver.find_elements(By.CSS_SELECTOR, 'h1, h2, h3, .title, [class*="title"]')
                    for heading in headings:
                        text = heading.text.strip()
                        if 'общее число запросов' in text.lower() or 'число запросов' in text.lower():
                            print(f"    Найден заголовок: {text}")
                            numbers = re.findall(r'\b(\d{1,3}(?:\s\d{3})*)\b', text)
                            if numbers:
                                frequency = int(numbers[-1].replace(' ', ''))
                                print(f"    Извлечена частота из заголовка: {frequency}")
                                break
                except Exception as e:
                    print(f"    Ошибка поиска в заголовках: {e}")
            
            # Метод 3: Поиск по XPath (альтернативный)
            if frequency is None:
                try:
                    # Ищем элементы, содержащие текст с числами и датами
                    xpath_selectors = [
                        "//div[contains(text(), ':')]",
                        "//span[contains(text(), ':')]",
                        "//*[contains(text(), 'число запросов')]",
                        "//*[contains(text(), '–') and contains(text(), ':')]"
                    ]
                    
                    for xpath in xpath_selectors:
                        elements = self.driver.find_elements(By.XPATH, xpath)
                        for element in elements:
                            text = element.text.strip()
                            if ':' in text:
                                print(f"    XPath найден: {text}")
                                # Ищем число после двоеточия
                                frequency_match = re.search(r':\s*(\d{1,3}(?:\s\d{3})*)', text)
                                if frequency_match:
                                    frequency = int(frequency_match.group(1).replace(' ', ''))
                                    print(f"    Извлечена частота по XPath: {frequency}")
                                    break
                        if frequency:
                            break
                except Exception as e:
                    print(f"    Ошибка XPath поиска: {e}")
            
            # Метод 4: Поиск в исходном коде страницы
            if frequency is None:
                try:
                    page_source = self.driver.page_source
                    print("    Ищем в исходном коде страницы...")
                    
                    # Ищем паттерны с датами и числами
                    patterns = [
                        r'за\s+\d{2}\.\d{2}\.\d{4}\s*–\s*\d{2}\.\d{2}\.\d{4}:\s*(\d{1,3}(?:\s\d{3})*)',
                        r'число запросов[^:]+:\s*(\d{1,3}(?:\s\d{3})*)',
                        r'общее число[^:]+:\s*(\d{1,3}(?:\s\d{3})*)',
                        r':\s*(\d{1,3}(?:\s\d{3})*)</div>'
                    ]
                    
                    for pattern in patterns:
                        matches = re.findall(pattern, page_source, re.IGNORECASE)
                        if matches:
                            frequency = int(matches[-1].replace(' ', ''))
                            print(f"    Найдена частота в исходном коде: {frequency}")
                            break
                    
                    # Если все еще не найдено, ищем любые большие числа
                    if frequency is None:
                        numbers = re.findall(r'\b(\d{1,3}(?:\s\d{3})+)\b', page_source)
                        numbers = [int(n.replace(' ', '')) for n in numbers if int(n.replace(' ', '')) > 100]
                        if numbers:
                            frequency = numbers[0]  # Берем первое большое число
                            print(f"    Найдено большое число как частота: {frequency}")
                            
                except Exception as e:
                    print(f"    Ошибка поиска в исходном коде: {e}")
            
            print(f"  Итоговая найденная частота: {frequency}")
            return frequency
            
        except Exception as e:
            print(f"  ✗ Ошибка парсинга для запроса '{query}': {e}")
            return None
    
    def parse_frequency_requests(self, query):
        """
        Парсинг частоты запроса с использованием requests
        
        Args:
            query (str): Поисковый запрос
            
        Returns:
            int or None: Частота запроса или None в случае ошибки
        """
        try:
            url = self.build_wordstat_url(query)
            print(f"  Запрос (requests): {query}")
            print(f"  URL: {url}")
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
                'Accept-Encoding': 'gzip, deflate',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1',
            }
            
            response = requests.get(url, headers=headers, timeout=15)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Метод 1: Поиск по CSS классам Яндекс Вордстат
            frequency_selectors = [
                '.wordstat__content-preview-text_last',
                '.wordstat__content-preview-text',
                '.wordstat__number',
                'div[class*="wordstat__"]',
                'span[class*="wordstat"]'
            ]
            
            frequency = None
            
            for selector in frequency_selectors:
                elements = soup.select(selector)
                for element in elements:
                    text = element.get_text().strip()
                    if text:
                        print(f"    Найден элемент '{selector}': {text}")
                        
                        # Ищем число после двоеточия (формат "за дата – дата: ЧИСЛО")
                        frequency_match = re.search(r':\s*(\d{1,3}(?:\s\d{3})*)', text)
                        if frequency_match:
                            frequency = int(frequency_match.group(1).replace(' ', ''))
                            print(f"    Извлечена частота: {frequency}")
                            break
                
                if frequency:
                    break
            
            # Метод 2: Поиск в HTML по регулярным выражениям
            if frequency is None:
                print("    Ищем в HTML коде...")
                html_content = response.text
                
                patterns = [
                    r'за\s+\d{2}\.\d{2}\.\d{4}\s*–\s*\d{2}\.\d{2}\.\d{4}:\s*(\d{1,3}(?:\s\d{3})*)',
                    r'число запросов[^:]+:\s*(\d{1,3}(?:\s\d{3})*)',
                    r'общее число[^:]+:\s*(\d{1,3}(?:\s\d{3})*)',
                    r'wordstat__content-preview-text[^>]*>([^<]*:\s*(\d{1,3}(?:\s\d{3})*))',
                    r'class="[^"]*wordstat[^"]*"[^>]*>([^<]*(\d{1,3}(?:\s\d{3})*))'
                ]
                
                for pattern in patterns:
                    matches = re.findall(pattern, html_content, re.IGNORECASE)
                    if matches:
                        # Извлекаем числа из найденных совпадений
                        for match in matches:
                            if isinstance(match, tuple):
                                # Берем последний элемент кортежа (обычно число)
                                number_str = match[-1]
                            else:
                                number_str = match
                            
                            if re.match(r'\d{1,3}(?:\s\d{3})*', number_str):
                                frequency = int(number_str.replace(' ', ''))
                                print(f"    Найдена частота в HTML: {frequency}")
                                break
                        if frequency:
                            break
            
            # Метод 3: Поиск любых больших чисел (последний вариант)
            if frequency is None:
                print("    Ищем любые большие числа...")
                # Ищем все числа в HTML
                all_numbers = re.findall(r'\b(\d{1,3}(?:\s\d{3})+)\b', response.text)
                if all_numbers:
                    # Конвертируем в int и фильтруем большие числа
                    numbers = [int(n.replace(' ', '')) for n in all_numbers]
                    numbers = [n for n in numbers if n > 1000]  # Исключаем маленькие числа
                    
                    if numbers:
                        frequency = numbers[0]  # Берем первое большое число
                        print(f"    Найдено большое число: {frequency}")
            
            print(f"  Итоговая найденная частота (requests): {frequency}")
            return frequency
            
        except Exception as e:
            print(f"  ✗ Ошибка requests для запроса '{query}': {e}")
            return None
    
    def get_query_frequency(self, query, query_type="base"):
        """
        Получение частоты запроса
        
        Args:
            query (str): Исходный запрос
            query_type (str): Тип запроса
            
        Returns:
            int or None: Частота запроса
        """
        formatted_query = self.format_query(query, query_type)
        
        if self.use_selenium:
            return self.parse_frequency_selenium(formatted_query)
        else:
            return self.parse_frequency_requests(formatted_query)
    
    def read_queries_from_file(self, filename):
        """
        Чтение запросов из файла
        
        Args:
            filename (str): Путь к файлу с запросами
            
        Returns:
            list: Список запросов
        """
        try:
            with open(filename, 'r', encoding='utf-8') as file:
                queries = [line.strip() for line in file if line.strip()]
            print(f"✓ Прочитано {len(queries)} запросов из файла {filename}")
            return queries
        except Exception as e:
            print(f"✗ Ошибка чтения файла {filename}: {e}")
            return []
    
    def create_excel_report(self, results, output_filename="wordstat_report.xlsx"):
        """
        Создание Excel отчета
        
        Args:
            results (list): Список результатов парсинга
            output_filename (str): Имя выходного файла
        """
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Анализ запросов Вордстат"
            
            # Заголовки колонок
            headers = ["Запрос", "Частота (базовая)", "Частота (точная)", "Частота (уточненная)"]
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
            
            # Заполнение данных
            for row_idx, result in enumerate(results, 2):
                query = result['query']
                base_freq = result.get('base_frequency', 'N/A')
                exact_freq = result.get('exact_frequency', 'N/A')
                precise_freq = result.get('precise_frequency', 'N/A')
                
                # Создаем гиперссылку для запроса
                url = self.build_wordstat_url(query)
                worksheet.cell(row=row_idx, column=1).hyperlink = url
                worksheet.cell(row=row_idx, column=1).value = query
                worksheet.cell(row=row_idx, column=1).font = Font(color="0000FF", underline="single")
                
                worksheet.cell(row=row_idx, column=2).value = base_freq
                worksheet.cell(row=row_idx, column=3).value = exact_freq
                worksheet.cell(row=row_idx, column=4).value = precise_freq
            
            # Автоматическая ширина колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(output_filename)
            print(f"✓ Excel отчет сохранен: {output_filename}")
            
        except Exception as e:
            print(f"✗ Ошибка создания Excel файла: {e}")
    
    def process_queries(self, queries):
        """
        Обработка списка запросов
        
        Args:
            queries (list): Список запросов для обработки
            
        Returns:
            list: Результаты парсинга
        """
        results = []
        total_queries = len(queries)
        
        print(f"\n🚀 Начинаю обработку {total_queries} запросов...")
        
        # Авторизация в Вордстат перед началом парсинга
        if self.use_selenium and self.driver and not self.is_authorized:
            print("\n" + "="*60)
            auth_success = self.authorize_wordstat()
            print("="*60)
            
            if not auth_success:
                print("⚠️  Продолжаем без авторизации (с медленными запросами)")
        
        # Определяем задержку между запросами
        if self.is_authorized:
            delay = 0.5  # Быстро если авторизован
            print(f"✅ Авторизован! Используем задержку {delay} сек между запросами")
        else:
            delay = 2.0  # Медленно если не авторизован
            print(f"⚠️  Не авторизован. Используем задержку {delay} сек между запросами")
        
        for idx, query in enumerate(queries, 1):
            print(f"\n[{idx}/{total_queries}] Обрабатываю: '{query}'")
            
            result = {'query': query}
            
            # Парсим базовую частоту
            print("  📊 Базовая частота...")
            result['base_frequency'] = self.get_query_frequency(query, "base")
            time.sleep(delay)  # Задержка между запросами
            
            # Парсим точную частоту
            print("  🎯 Точная частота...")
            result['exact_frequency'] = self.get_query_frequency(query, "exact")
            time.sleep(delay)
            
            # Парсим уточненную частоту
            print("  🔍 Уточненная частота...")
            result['precise_frequency'] = self.get_query_frequency(query, "precise")
            time.sleep(delay)
            
            results.append(result)
            
            print(f"  ✓ Результат: {result['base_frequency']} | {result['exact_frequency']} | {result['precise_frequency']}")
        
        return results
    
    def close(self):
        """Закрытие WebDriver"""
        if self.driver:
            self.driver.quit()
            print("✓ WebDriver закрыт")


def main():
    """Основная функция программы"""
    print("=== Парсер Яндекс Вордстат ===\n")
    
    # Проверяем наличие файла с запросами
    input_file = "queries.txt"
    if not os.path.exists(input_file):
        print(f"✗ Файл {input_file} не найден!")
        print(f"Создайте файл {input_file} и добавьте в него запросы (по одному на строку)")
        return
    
    # Создаем экземпляр парсера
    parser = WordstatParser(use_selenium=True)
    
    try:
        # Читаем запросы из файла
        queries = parser.read_queries_from_file(input_file)
        if not queries:
            print("✗ Не удалось прочитать запросы из файла")
            return
        
        # Обрабатываем запросы
        results = parser.process_queries(queries)
        
        # Создаем Excel отчет
        output_file = "wordstat_report.xlsx"
        parser.create_excel_report(results, output_file)
        
        print(f"\n🎉 Готово! Результаты сохранены в {output_file}")
        
    except KeyboardInterrupt:
        print("\n⚠️  Работа прервана пользователем")
    except Exception as e:
        print(f"\n✗ Произошла ошибка: {e}")
    finally:
        parser.close()


if __name__ == "__main__":
    main()
